import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Content Calendar"

# Define headers
headers = ["Date", "Day", "Video Title", "Topic/Niche", "Keywords", "Status", "Publish Time", "Notes", "Thumbnail Idea", "CTA"]

# Add headers to the sheet and apply styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Set column widths
sheet.column_dimensions["A"].width = 12
sheet.column_dimensions["B"].width = 12
sheet.column_dimensions["C"].width = 40
sheet.column_dimensions["D"].width = 20
sheet.column_dimensions["E"].width = 40
sheet.column_dimensions["F"].width = 15
sheet.column_dimensions["G"].width = 15
sheet.column_dimensions["H"].width = 40
sheet.column_dimensions["I"].width = 40
sheet.column_dimensions["J"].width = 30

# Add example data for one week
from datetime import date, timedelta

today = date.today()
for i in range(7):
    current_date = today + timedelta(days=i)
    row_num = i + 2
    sheet.cell(row=row_num, column=1, value=current_date.strftime("%Y-%m-%d"))
    sheet.cell(row=row_num, column=2, value=current_date.strftime("%A"))
    sheet.cell(row=row_num, column=3, value=f"Daily Insights: [Topic of the Day]")
    sheet.cell(row=row_num, column=4, value="[Your Niche]")
    sheet.cell(row=row_num, column=5, value="[Keyword 1], [Keyword 2], [Keyword 3]")
    sheet.cell(row=row_num, column=6, value="Planned")
    sheet.cell(row=row_num, column=7, value="09:00 AM EST")
    sheet.cell(row=row_num, column=8, value="Source: [Link to research report]")
    sheet.cell(row=row_num, column=9, value="[Idea for thumbnail, e.g., 'Surprised face with question mark']")
    sheet.cell(row=row_num, column=10, value="Subscribe, Like, Comment")

# Apply cell formatting
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Save the workbook
workbook.save("/home/ubuntu/templates/calendars/content_calendar_template.xlsx")

print("Content calendar template created successfully.")

